import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import requests
from datetime import datetime, timedelta
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import os
import time
import json
from typing import Dict, Optional
import yfinance as yf
from bs4 import BeautifulSoup
import logging

# Set up logging
logging.basicConfig(
    filename='price_tracker.log',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

class PriceRatioApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Cryptocurrency to Gold Price Ratio")
        self.root.geometry("1000x800")

        # Initialize variables
        self.days_to_show = tk.IntVar(value=7)
        self.ratio_mode = tk.StringVar(value="btc_gold")
        self.current_data = []
        self.api_status = {}

        # Create main frame
        self.main_frame = ttk.Frame(self.root, padding="10")
        self.main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        # Create control frame
        self.create_control_frame()

        # Create table frame
        self.create_table_frame()

        # Create status frame
        self.create_status_frame()

        # Initial data load
        self.refresh_data()

    def create_status_frame(self):
        """Create frame for API status information"""
        self.status_frame = ttk.LabelFrame(self.main_frame, text="API Status", padding="5")
        self.status_frame.grid(row=2, column=0, pady=5, sticky="ew")

        self.status_labels = {}
        for api_name in ["NBP", "YFinance", "CoinGecko"]:
            self.status_labels[api_name] = ttk.Label(self.status_frame, text=f"{api_name}: Waiting...")
            self.status_labels[api_name].pack(anchor="w")

    def update_api_status(self, api_name: str, status: bool, message: str = ""):
        """Update API status display with more detail"""
        status_text = "✓ Success" if status else f"✗ Failed: {message}"
        color = "green" if status else "red"
        timestamp = datetime.now().strftime('%H:%M:%S')
        self.status_labels[api_name].config(
            text=f"{api_name} [{timestamp}]: {status_text}",
            foreground=color
        )
        self.api_status[api_name] = status
        logging.info(f"API Status Update - {api_name}: {status_text}")

    def create_control_frame(self):
        """Create frame for controls"""
        control_frame = ttk.LabelFrame(self.main_frame, text="Controls", padding="5")
        control_frame.grid(row=0, column=0, pady=5, sticky="ew")

        # Days selection
        days_frame = ttk.Frame(control_frame)
        days_frame.pack(fill="x", pady=5)

        ttk.Label(days_frame, text="Number of days to show:").pack(side=tk.LEFT, padx=5)
        days_entry = ttk.Entry(days_frame, textvariable=self.days_to_show, width=5)
        days_entry.pack(side=tk.LEFT, padx=5)

        # Preset buttons for days
        presets_frame = ttk.Frame(control_frame)
        presets_frame.pack(fill="x", pady=5)

        for days in [7, 14, 30, 90]:
            ttk.Button(presets_frame,
                       text=f"{days}d",
                       command=lambda d=days: self.set_days(d)).pack(side=tk.LEFT, padx=2)

        # Ratio selection buttons
        ratio_frame = ttk.Frame(control_frame)
        ratio_frame.pack(fill="x", pady=5)

        ttk.Radiobutton(ratio_frame,
                        text="BTC/Gold Ratio",
                        variable=self.ratio_mode,
                        value="btc_gold",
                        command=self.refresh_data).pack(side=tk.LEFT, padx=5)

        ttk.Radiobutton(ratio_frame,
                        text="Gold/BTC Ratio",
                        variable=self.ratio_mode,
                        value="gold_btc",
                        command=self.refresh_data).pack(side=tk.LEFT, padx=5)

        # Refresh and Export buttons
        button_frame = ttk.Frame(control_frame)
        button_frame.pack(fill="x", pady=5)

        ttk.Button(button_frame,
                   text="Refresh Data",
                   command=self.refresh_data).pack(side=tk.LEFT, padx=5)

        ttk.Button(button_frame,
                   text="Export to Excel",
                   command=self.export_to_excel).pack(side=tk.LEFT, padx=5)

        # Last refresh label
        self.refresh_label = ttk.Label(control_frame, text="")
        self.refresh_label.pack(fill="x", pady=5)

    def create_table_frame(self):
        """Create frame for data table"""
        table_frame = ttk.LabelFrame(self.main_frame, text="Price Data", padding="5")
        table_frame.grid(row=1, column=0, pady=5, sticky="nsew")

        # Configure columns
        columns = ("Date", "Gold Price (USD)", "Bitcoin Price (USD)", "Ratio")
        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings")

        # Add scrollbar
        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        scrollbar.pack(side=tk.RIGHT, fill="y")
        self.tree.configure(yscrollcommand=scrollbar.set)

        # Configure columns
        self.tree.column("Date", width=120)
        self.tree.column("Gold Price (USD)", width=150)
        self.tree.column("Bitcoin Price (USD)", width=150)
        self.tree.column("Ratio", width=150)

        # Configure headings
        for col in columns:
            self.tree.heading(col, text=col)

        self.tree.pack(fill="both", expand=True)

    def get_gold_prices_historical(self) -> Optional[Dict[str, float]]:
        """Fetch historical gold prices with fallback mechanism"""
        try:
            # First try to get from local cache
            cached_prices = self.load_cached_prices()
            if cached_prices:
                self.update_api_status("NBP", True, "Using cached data")
                return cached_prices

            # Try NBP API
            url = "https://api.nbp.pl/api/cenyzlota/last/{}".format(self.days_to_show.get())
            headers = {
                'Accept': 'application/json'
            }

            response = requests.get(url, headers=headers)
            if response.status_code == 200:
                data = response.json()
                prices = {}
                for item in data:
                    date = item['data']
                    # Convert from PLN/g to USD/oz (approximate conversion)
                    price_pln_per_gram = float(item['cena'])
                    price_usd_per_oz = price_pln_per_gram * 31.1034768 * 0.25  # Approximate PLN to USD conversion
                    prices[date] = price_usd_per_oz

                # Cache the results
                self.cache_prices(prices)
                self.update_api_status("NBP", True)
                return prices

            self.update_api_status("NBP", False, f"Status code: {response.status_code}")
            return self.get_gold_prices_yfinance()

        except Exception as e:
            logging.error(f"Error fetching NBP gold prices: {str(e)}")
            self.update_api_status("NBP", False, str(e))
            return self.get_gold_prices_yfinance()

    def get_gold_prices_yfinance(self) -> Optional[Dict[str, float]]:
        """Fetch gold prices using yfinance as fallback"""
        try:
            # Get GLD (SPDR Gold Shares) data as a proxy for gold prices
            gld = yf.Ticker("GLD")
            end_date = datetime.now()
            start_date = end_date - timedelta(days=self.days_to_show.get())

            df = gld.history(start=start_date, end=end_date)

            prices = {}
            for date, row in df.iterrows():
                date_str = date.strftime('%Y-%m-%d')
                # Convert GLD price to approximate gold price
                prices[date_str] = float(row['Close']) * 10

            self.update_api_status("YFinance", True)
            return prices

        except Exception as e:
            logging.error(f"Error fetching YFinance gold prices: {str(e)}")
            self.update_api_status("YFinance", False, str(e))
            return None

    def get_bitcoin_prices_historical(self) -> Optional[Dict[str, float]]:
        """Fetch historical bitcoin prices from CoinGecko"""
        try:
            end_date = datetime.now()
            start_date = end_date - timedelta(days=self.days_to_show.get())

            url = 'https://api.coingecko.com/api/v3/coins/bitcoin/market_chart/range'
            params = {
                'vs_currency': 'usd',
                'from': int(start_date.timestamp()),
                'to': int(end_date.timestamp())
            }

            response = requests.get(url, params=params)
            if response.status_code == 200:
                data = response.json()
                prices = {}
                for timestamp, price in data['prices']:
                    date = datetime.fromtimestamp(timestamp/1000).strftime('%Y-%m-%d')
                    if date not in prices:  # Take the last price of the day
                        prices[date] = price
                self.update_api_status("CoinGecko", True)
                return prices

            self.update_api_status("CoinGecko", False, f"Status code: {response.status_code}")
            return None
        except Exception as e:
            logging.error(f"Error fetching Bitcoin prices: {str(e)}")
            self.update_api_status("CoinGecko", False, str(e))
            return None

    def cache_prices(self, prices: Dict[str, float]):
        """Cache prices locally"""
        try:
            with open('gold_prices_cache.json', 'w') as f:
                json.dump({
                    'timestamp': datetime.now().timestamp(),
                    'prices': prices
                }, f)
        except Exception as e:
            logging.error(f"Error caching prices: {str(e)}")

    def load_cached_prices(self) -> Optional[Dict[str, float]]:
        """Load cached prices if they're recent enough"""
        try:
            with open('gold_prices_cache.json', 'r') as f:
                data = json.load(f)
                cache_time = datetime.fromtimestamp(data['timestamp'])
                if datetime.now() - cache_time < timedelta(hours=1):  # Cache valid for 1 hour
                    return data['prices']
        except Exception:
            return None

    def calculate_ratio(self, gold_price: float, btc_price: float) -> float:
        """Calculate ratio based on selected mode"""
        if self.ratio_mode.get() == "btc_gold":
            return btc_price / gold_price
        else:
            return gold_price / btc_price

    def get_ratio_column_name(self) -> str:
        """Get the appropriate ratio column name"""
        return "BTC/Gold Ratio" if self.ratio_mode.get() == "btc_gold" else "Gold/BTC Ratio"

    def set_days(self, days: int):
        """Set the number of days and refresh data"""
        self.days_to_show.set(days)
        self.refresh_data()

    def refresh_data(self):
        """Refresh all data"""
        try:
            days = int(self.days_to_show.get())
            if days <= 0:
                messagebox.showerror("Error", "Please enter a positive number of days")
                return
        except ValueError:
            messagebox.showerror("Error", "Please enter a valid number")
            return

        # Clear existing items
        for item in self.tree.get_children():
            self.tree.delete(item)

        # Clear current data
        self.current_data = []

        # Update ratio column heading
        self.tree.heading("Ratio", text=self.get_ratio_column_name())

        # Get historical prices
        gold_prices = self.get_gold_prices_historical()
        bitcoin_prices = self.get_bitcoin_prices_historical()

        if gold_prices and bitcoin_prices:
            dates = sorted(set(gold_prices.keys()) & set(bitcoin_prices.keys()))[-self.days_to_show.get():]

            for date in dates:
                gold_price = gold_prices.get(date)
                bitcoin_price = bitcoin_prices.get(date)

                if gold_price and bitcoin_price:
                    ratio = self.calculate_ratio(gold_price, bitcoin_price)

                    # Store data for export
                    self.current_data.append({
                        'Date': date,
                        'Gold Price (USD)': gold_price,
                        'Bitcoin Price (USD)': bitcoin_price,
                        'Ratio': ratio
                    })

                    # Insert data into treeview
                    self.tree.insert("", "end", values=(
                        date,
                        f"${gold_price:,.2f}",
                        f"${bitcoin_price:,.2f}",
                        f"{ratio:.6f}"
                    ))

            # Update refresh time label
            self.refresh_label.config(
                text=f"Last refreshed: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            )
        else:
            self.refresh_label.config(text="Error fetching data. Please try again.")
            messagebox.showerror("Error", "Failed to fetch price data")

    def export_to_excel(self):
        """Export data to Excel file"""
        if not self.current_data:
            messagebox.showerror("Error", "No data to export!")
            return

        try:
            # Ask user for save location
            file_path = filedialog.asksaveasfilename(
                defaultextension='.xlsx',
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=f"price_ratio_{self.days_to_show.get()}days_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )

            if not file_path:  # If user cancels the file dialog
                return

            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Price Ratio Data"

            # Define headers
            headers = ["Date", "Gold Price (USD)", "Bitcoin Price (USD)", self.get_ratio_column_name()]

            # Style for headers
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

            # Write headers and style them
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill

            # Write data
            for row, data in enumerate(self.current_data, 2):
                ws.cell(row=row, column=1, value=data['Date'])
                ws.cell(row=row, column=2, value=data['Gold Price (USD)'])
                ws.cell(row=row, column=3, value=data['Bitcoin Price (USD)'])
                ws.cell(row=row, column=4, value=data['Ratio'])

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width

            # Save the file
            wb.save(file_path)
            messagebox.showinfo("Success", f"Data exported successfully to:\n{file_path}")

        except Exception as e:
            logging.error(f"Error exporting to Excel: {str(e)}")
            messagebox.showerror("Error", f"An error occurred while exporting:\n{str(e)}")

def main():
    root = tk.Tk()
    app = PriceRatioApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()